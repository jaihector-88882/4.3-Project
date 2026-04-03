print("<Jaihea8080>")

import csv
import os
import openpyxl
from openpyxl.chart import Reference, BarChart, LineChart
from datetime import datetime

path = "C:\\PythonFiles\\ZooData.csv"

# Inserting data into a csv file
def insertData(path, data):
    try:
        with open(path, 'a', newline='') as csvfile:
            csvfile.write(data + '\n')
    except Exception as e:
        print("Error writing to file:", e)

def viewData(path):
    print("The file", path)
    try:
        with open(path, 'r') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                print(','.join(row))
    except Exception as e:
        print("Error reading file:", e)

def getInput():
    dates = []
    temps = []

    while True:
        date = input("Enter a date: ")
        if date == "":
            break
        dates.append(date)

        temp = input("Enter the highest temp for the inputted date: ")
        if temp == "":
            break
        temps.append(int(temp))

    for i in range(len(dates)):
        avg = sum(temps) / len(temps)
        data = dates[i] + "," + str(temps[i]) + "," + str(avg)
        try:
            insertData(path, data)
            print("The following data was saved at " + str(datetime.now()) + ":")
            print(data)
        except Exception as e:
            print("Error saving data:", e)

def createChart(path, chartType):
    print("Which data source would you like to use?")
    print("1 Original Data (Fahrenheit)")
    print("2 Converted Data (Celsius)")
    dataChoice = input()

    dates = []
    originalData = []
    convertedData = []

    try:
        with open(path, 'r') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                if len(row) >= 3:
                    dates.append(row[0])
                    originalData.append(int(row[1]))
                    convertedData.append(float(row[2]))
    except Exception as e:
        print("Error reading file:", e)
        return

   
    wb = openpyxl.Workbook()
    ws = wb.active

   
    ws['A1'] = 'Date'
    ws['B1'] = 'Fahrenheit'
    ws['C1'] = 'Celsius'

   
    for i in range(len(dates)):
        ws['A' + str(i+2)] = dates[i]
        ws['B' + str(i+2)] = originalData[i]
        ws['C' + str(i+2)] = convertedData[i]

   
    if dataChoice == "1":
        dataCol = 2
        yLabel = "Fahrenheit"
    else:
        dataCol = 3
        yLabel = "Celsius"

   
    chartTitle = "Jaihea8080 " + datetime.now().strftime("%m/%d/%Y")

    
    if chartType == "bar":
        myChart = BarChart()
    else:
        myChart = LineChart()

    myChart.title = chartTitle
    myChart.legend = None

  
    myData = Reference(ws, min_col=dataCol, min_row=1, max_row=len(dates)+1)
    myLabels = Reference(ws, min_col=1, min_row=2, max_row=len(dates)+1)

    myChart.add_data(myData, titles_from_data=True)
    myChart.set_categories(myLabels)

    myChart.x_axis.title = "Date"
    myChart.y_axis.title = yLabel

    
    ws.add_chart(myChart, "E2")

    
    wb.save("C:\\PythonFiles\\final.xlsx")
    print("Chart saved to C:\\PythonFiles\\final.xlsx")


def generateReport(path):
    print("Which graph type would you like to create?")
    print("1 Bar Chart")
    print("2 Line Chart")
    graphChoice = input()

    if graphChoice == "1":
        createChart(path, "bar")
    elif graphChoice == "2":
        createChart(path, "line")
    else:
        print("Invalid selection")

def main():
    print("=" * 60)
    print("Jaihea8080's Spreadsheet Automation Menu")
    print("Choose a number from the following options")
    print("1 Input Data")
    print("2 View Current Data")
    print("3 Generate Report")

    choice = input()
    print("You selected " + choice + " at " + str(datetime.now()))

    if choice == "1":
        getInput()
    elif choice == "2":
        viewData(path)
    elif choice == "3":
        generateReport(path)
    else:
        print("Invalid selection")

main()


